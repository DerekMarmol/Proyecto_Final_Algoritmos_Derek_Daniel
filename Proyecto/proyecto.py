import argparse
import openpyxl
import sys

def Ayuda():
    print("Puede utilizar los siguientes comandos según lo que desee hacer:")
    print("Ver la lista de productos: python Proyecto_Final.py --inventario_listar")
    print("Crear un producto: python Proyecto_Final.py --inventario_crear")
    print("Actualizar un producto: python Proyecto_Final.py --inventario_actualizar")

#Modulo de Inventario

def listar_producto():
    book = openpyxl.load_workbook('inventario.xlsx')
    sheet = book['Inventario']

    headers = [cell.value for cell in sheet[1]]

    data_by_header = {header: [] for header in headers}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        for header, cell_value in zip(headers, row):
            data_by_header[header].append(cell_value)

    max_data_length = max(len(data) for data in data_by_header.values())

    for header in headers:
        print(f"{header.ljust(12)}", end='')
    print()  

    for i in range(max_data_length):
        for header, data in data_by_header.items():
            if i < len(data):
                print(f"{str(data[i]).ljust(12)}", end='')
            else:
                print("".ljust(12), end='')  
        print() 

def crear_producto():
    book = openpyxl.load_workbook('inventario.xlsx')
    sheet = book['Inventario']

    producto = int(input("Ingrese el código del producto: "))
    nombre = input("Ingrese el nombre del producto: ")
    existencia = int(input("Ingrese la existencia del producto: "))
    proveedor = input("Ingrese el proveedor del producto: ")
    precio = input("Ingrese el precio del producto: ")

    data = [producto, nombre, existencia, proveedor, precio]
    sheet.append(data)

    book.save('inventario.xlsx')

def actualizar_producto():
    try:
        book = openpyxl.load_workbook('inventario.xlsx')
        sheet = book['Inventario']

        codigo = int(input("Ingrese el código del producto que desea actualizar: "))

        producto_encontrado = False

        for fila in sheet.iter_rows(min_col=1, max_col=1):
            celda = fila[0]
            if celda.value == codigo:
                print("Se ha encontrado el producto")
                producto_encontrado = True
                fila_destino = celda.row 
                columna_destino = 5
                celda_destino = sheet.cell(row=fila_destino, column=columna_destino)
                precio_nuevo = input("Ingrese el nuevo precio: ")
                celda_destino.value = precio_nuevo
                break 

        if not producto_encontrado:
            print("No se encontró el producto con el código proporcionado")

        book.save('inventario.xlsx')

    except Exception as e:
        print("Se produjo un error:", str(e))

def actualizar_existencia():
    try:
        book = openpyxl.load_workbook('inventario.xlsx')
        sheet = book['Inventario']

        codigo = int(input("Ingrese el código del producto que desea actualizar: "))

        producto_encontrado = False

        for fila in sheet.iter_rows(min_col=1, max_col=1):
            celda = fila[0]
            if celda.value == codigo:
                print("Se ha encontrado el producto")
                producto_encontrado = True
                fila_destino = celda.row 
                columna_destino = 3
                celda_destino = sheet.cell(row=fila_destino, column=columna_destino)
                precio_nuevo = input("Ingrese la nueva existencia: ")
                celda_destino.value = precio_nuevo
                break 

        if not producto_encontrado:
            print("No se encontró el producto con el código proporcionado")

        book.save('inventario.xlsx')

    except Exception as e:
        print("Se produjo un error:", str(e))

def eliminar_producto():
    try:
        book = openpyxl.load_workbook('inventario.xlsx')
        sheet = book['Inventario']

        codigo = int(input("Ingrese el código del producto que desea eliminar: "))

        producto_encontrado = False

        for fila in sheet.iter_rows(min_col=1, max_col=1):
            celda = fila[0]
            if celda.value == codigo:
                print("Se ha encontrado el producto")
                producto_encontrado = True
                fila_destino = celda.row 
                sheet.delete_rows(fila_destino)
                print("El producto ha sido eliminado.")
                break 

        if not producto_encontrado:
            print("No se encontró el producto con el código proporcionado")

        book.save('inventario.xlsx')

    except Exception as e:
        print("Se produjo un error:", str(e))



#Modulo de clientes

def listar_clientes():
    book = openpyxl.load_workbook('inventario.xlsx')
    sheet = book['Clientes']

    headers = [cell.value for cell in sheet[1]]

    data_by_header = {header: [] for header in headers}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        for header, cell_value in zip(headers, row):
            data_by_header[header].append(cell_value)

    max_data_length = max(len(data) for data in data_by_header.values())

    for header in headers:
        print(f"{header.ljust(12)}", end='')
    print()  

    for i in range(max_data_length):
        for header, data in data_by_header.items():
            if i < len(data):
                print(f"{str(data[i]).ljust(12)}", end='')
            else:
                print("".ljust(12), end='')  
        print()

def crear_clientes():
    book = openpyxl.load_workbook('inventario.xlsx')
    sheet = book['Clientes']

    while True:
        codigo = int(input("Ingrese el código del cliente a agregar (o -1 para salir): "))
        
        if codigo == -1:
            break

        nombre = input("Ingrese el nombre del cliente: ")
        Direccion = input("Ingrese la dirección del cliente: ")

        clientes_existentes = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            codigo_cliente, nombre_cliente, direccion_cliente = row
            clientes_existentes.append((codigo_cliente, nombre_cliente, direccion_cliente))

        if any(cliente[0] == codigo for cliente in clientes_existentes):
            print(f"El cliente con código {codigo} ya existe. Introduzca otro código.")
        else:
            nuevo_cliente = (codigo, nombre, Direccion)
            clientes_existentes.append(nuevo_cliente)

            clientes_ordenados = sorted(clientes_existentes, key=lambda cliente: cliente[0])

            for _ in range(2, sheet.max_row + 1):
                sheet.delete_rows(2)

            for cliente in clientes_ordenados:
                sheet.append(cliente)

            print("El nuevo cliente fue agregado de forma correcta")

            book.save('inventario.xlsx')
            break

def actualizar_cliente():
    try:
        book = openpyxl.load_workbook('inventario.xlsx')
        sheet = book['Clientes']

        codigo = int(input("Ingrese el código del cliente que desea actualizar: "))

        cliente_encontrado = False

        for fila in sheet.iter_rows(min_col=1, max_col=1):
            celda = fila[0]
            if celda.value == codigo:
                print("Se ha encontrado el cliente")
                cliente_encontrado = True
                fila_destino = celda.row 
                columna_destino = 3
                celda_destino = sheet.cell(row=fila_destino, column=columna_destino)
                direccion_nueva = input("Ingrese la nueva dirección: ")
                celda_destino.value = direccion_nueva
                print("Los datos del cliente se actualizaron correctamente")
                break 

        if not cliente_encontrado:
            print("No se encontró el cliente con el código proporcionado")

        book.save('inventario.xlsx')

    except Exception as e:
        print("Se produjo un error:", str(e))

def eliminar_cliente():
    try:
        book = openpyxl.load_workbook('inventario.xlsx')
        sheet = book['Clientes']

        codigo = int(input("Ingrese el código del cliente que desea eliminar: "))

        cliente_encontrado = False

        for fila in sheet.iter_rows(min_col=1, max_col=1):
            celda = fila[0]
            if celda.value == codigo:
                print("Se ha encontrado el cliente")
                cliente_encontrado = True
                fila_destino = celda.row 
                sheet.delete_rows(fila_destino) 
                print("El cliente ha sido eliminado.")
                break 

        if not cliente_encontrado:
            print("No se encontró el cliente con el código proporcionado")

        book.save('inventario.xlsx')

    except Exception as e:
        print("Se produjo un error:", str(e))

#Modulo de Control de Ventas

def listar_ventas():
    book = openpyxl.load_workbook('inventario.xlsx')
    sheet = book['Ventas']

    headers = [cell.value for cell in sheet[1]]

    data_by_header = {header: [] for header in headers}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        for header, cell_value in zip(headers, row):
            data_by_header[header].append(cell_value)

    max_data_length = max(len(data) for data in data_by_header.values())

    for header in headers:
        print(f"{header.ljust(20)}", end='')
    print()

    for i in range(max_data_length):
        for header, data in data_by_header.items():
            if i < len(data):
                value = str(data[i])
                if header == "Total Venta":
                    print(f"{value.ljust(20)}", end='')
                else:
                    print(f"{value.ljust(20)}", end='')
            else:
                print("".ljust(20), end='')
        print()


def verificar_producto_en_inventario(codigo_producto, hoja_inventario):
    for row in hoja_inventario.iter_rows(values_only=True):
        if row[0] == codigo_producto:
            return row

    return None

def agregar_venta():
    codigo_producto = int(input("Ingrese el código del producto a vender: "))
    codigo_cliente = int(input("Ingrese el código del cliente: "))

    book = openpyxl.load_workbook('inventario.xlsx')
    
    hoja_inventario = book['Inventario']

    producto = verificar_producto_en_inventario(codigo_producto, hoja_inventario)

    if producto is None:
        print(f"El producto con código {codigo_producto} no se encuentra en el inventario.")
        return

    cantidad_existente = producto[2] 
    precio_unitario = producto[4]

    if cantidad_existente <= 0:
        print("El producto está agotado y no se puede vender.")
        return

    cantidad_vendida = int(input(f"Ingrese la cantidad a vender (existencia actual: {cantidad_existente}): "))

    if cantidad_vendida > cantidad_existente:
        print("No hay suficiente cantidad en inventario para la venta.")
        return

    total_venta = cantidad_vendida * precio_unitario

    hoja_ventas = book['Ventas']
    hoja_ventas.append([codigo_producto, codigo_cliente, cantidad_vendida, total_venta])

    for idx, row in enumerate(hoja_inventario.iter_rows(values_only=True, min_row=2), start=2):
        if row[0] == codigo_producto:
            cantidad_actual = row[2]
            nueva_cantidad = cantidad_actual - cantidad_vendida
            hoja_inventario.cell(row=idx, column=3, value=nueva_cantidad)

    book.save('inventario.xlsx')

    print(f"Venta registrada exitosamente. Total de la venta: {total_venta}.")

def anular_venta():
    book = openpyxl.load_workbook('inventario.xlsx')
    sheet = book['Ventas']

    codigo = int(input("Ingrese el codigo del producto vendido, por medio de este se eliminará la venta: "))

    venta_encontrada = False

    for fila in sheet.iter_rows(min_col=1, max_row=1):
        celda = fila[0]
        if celda.value == codigo:
                print("Se ha encontrado la venta")
                venta_encontrada = True
                fila_destino = celda.row 
                sheet.delete_rows(fila_destino) 
                print("la venta fue anulada/eliminada.")
                break 

        if not venta_encontrada:
            print("No se encontró la venta con el código proporcionado")

        book.save('inventario.xlsx')

#Modulo de consultas

def generar_informe_ventas_por_cliente(archivo_excel, hoja_ventas, nombre_archivo_salida):
    book = openpyxl.load_workbook(archivo_excel)
    sheet = book[hoja_ventas]

    ventas_por_cliente = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        codigo_producto, codigo_cliente, cantidad_producto, total_venta = row

        codigo_cliente_str = str(codigo_cliente)

        total_gastado = total_venta

        if codigo_cliente_str in ventas_por_cliente:
            ventas_por_cliente[codigo_cliente_str] += total_gastado
        else:
            ventas_por_cliente[codigo_cliente_str] = total_gastado

    nuevo_book = openpyxl.Workbook()
    nueva_sheet = nuevo_book.active

    nueva_sheet['A1'] = "Código del Cliente"
    nueva_sheet['B1'] = "Total Gastado"

    fila = 2
    for codigo_cliente, total_gastado in ventas_por_cliente.items():
        nueva_sheet[f'A{fila}'] = int(codigo_cliente) 
        nueva_sheet[f'B{fila}'] = total_gastado
        fila += 1

    nuevo_book.save(nombre_archivo_salida)

    

def generar_informe_ventas_por_producto(archivo_excel, hoja_ventas, nombre_archivo_salida):
    try:
        book = openpyxl.load_workbook('inventario.xlsx')
        sheet = book['Ventas']

        ventas_por_producto = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            codigo_producto, _, cantidad_producto, total_venta = row 

            
            if codigo_producto in ventas_por_producto:
                ventas_por_producto[codigo_producto]["Cantidad Vendida"] += cantidad_producto
                ventas_por_producto[codigo_producto]["Total Ventas"] += total_venta
            else:
               
                ventas_por_producto[codigo_producto] = {
                    "Cantidad Vendida": cantidad_producto,
                    "Total Ventas": total_venta
                }

        informe_book = openpyxl.Workbook()
        informe_sheet = informe_book.active
        informe_sheet.title = "Informe Ventas por Producto"

        informe_sheet.append(["Código de Producto", "Cantidad Total Vendida", "Total Ventas"])

        for codigo_producto, datos in ventas_por_producto.items():
            informe_sheet.append([codigo_producto, datos["Cantidad Vendida"], datos["Total Ventas"]])

        informe_book.save(nombre_archivo_salida)
        print(f"Informe de ventas por producto generado en '{nombre_archivo_salida}'.")

    except Exception as e:
        print("Se produjo un error:", str(e)) 
    
def menu_interactivo():
    while True:
        print("Menú de opciones:")
        print("1. Listar productos")
        print("2. Crear producto")
        print("3. Actualizar producto")
        print("4. Actualizar existencia")
        print("5. Eliminar producto")
        print("6. Listar clientes")
        print("7. Crear cliente")
        print("8. Actualizar cliente")
        print("9. Eliminar cliente")
        print("10. Listar ventas")
        print("11. Agregar venta")
        print("12. Anular venta")
        print("13. Generar informe de ventas por cliente.")
        print("14. Generar informe de ventas por producto.") 
        print("15. Salir")
        
        opcion = input("Ingrese el número de la opción que desee: ")
        
        if opcion == '1':
            listar_producto()
        elif opcion == '2':
            crear_producto()
        elif opcion == '3':
            actualizar_producto()
        elif opcion == '4':
            actualizar_existencia()
        elif opcion == '5':
            eliminar_producto()
        elif opcion == '6':
            listar_clientes()
        elif opcion == '7':
            crear_clientes()
        elif opcion == '8':
            actualizar_cliente()
        elif opcion == '9':
            eliminar_cliente()
        elif opcion == '10':
            listar_ventas()
        elif opcion == '11':
            agregar_venta()
        elif opcion == '12':
            anular_venta()
        elif opcion == '13':
            generar_informe_ventas_por_cliente('inventario.xlsx', 'Ventas', 'informe_ventas_por_cliente.xlsx')
        elif opcion == '14':
            generar_informe_ventas_por_producto('inventario.xlsx', 'Ventas', 'informe_ventas_por_producto.xlsx')  # Llama a la función de informe de ventas por producto
        elif opcion == '15':
            break
        else:
            print("Opción no válida. Intente nuevamente.")

        continuar = input("¿Desea ejecutar otra función? (S/N): ")
        if continuar.lower() != 's':
            break

def linea_de_comandos():
    parser = argparse.ArgumentParser()
    parser.add_argument('--inventario_listar', action='store_true')
    parser.add_argument('--ayuda', action='store_true')
    parser.add_argument('--inventario_crear', action='store_true')
    parser.add_argument('--inventario_actualizar', action='store_true')
    parser.add_argument('--inventario_existencia', action='store_true')
    parser.add_argument('--inventario_eliminar', action='store_true')
    parser.add_argument('--listar_clientes', action='store_true')
    parser.add_argument('--crear_clientes', action='store_true')
    parser.add_argument('--actualizar_cliente', action='store_true')
    parser.add_argument('--eliminar_clientes', action='store_true')
    parser.add_argument('--listar_ventas', action='store_true')
    parser.add_argument('--agregar_venta', action='store_true')
    parser.add_argument('--anular_ventas', action='store_true')
    args = parser.parse_args()

    if args.ayuda:
        Ayuda()
    elif args.inventario_listar:
        listar_producto()
    elif args.inventario_crear:
        crear_producto()
    elif args.inventario_actualizar:
        actualizar_producto()
    elif args.inventario_existencia:
        actualizar_existencia()
    elif args.inventario_eliminar:
        eliminar_producto()
    elif args.listar_clientes:
        listar_clientes()
    elif args.crear_clientes:
        crear_clientes()
    elif args.actualizar_cliente:
        actualizar_cliente()
    elif args.eliminar_clientes:
        eliminar_cliente()
    elif args.listar_ventas:
        listar_ventas()
    elif args.agregar_venta:
        agregar_venta()
    elif args.anular_ventas:
        anular_venta()

if __name__ == "__main__":
    if len(sys.argv) == 1:
        menu_interactivo()
    else:
        linea_de_comandos()