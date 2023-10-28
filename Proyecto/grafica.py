import openpyxl
import tkinter as tk
from tkinter import messagebox, simpledialog
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from tkinter import filedialog, simpledialog

#Modulo de Inventario
def listar_producto(text_widget):

    text_widget.delete('1.0', tk.END)

    book = openpyxl.load_workbook('inventario.xlsx')
    sheet = book['Inventario']

    headers = [cell.value for cell in sheet[1]]

    data_by_header = {header: [] for header in headers}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        for header, cell_value in zip(headers, row):
            data_by_header[header].append(cell_value)

    max_data_length = max(len(data) for data in data_by_header.values())

    for header in headers:
        text_widget.insert(tk.END, f"{header.ljust(12)}")
    text_widget.insert(tk.END, '\n')

    for i in range(max_data_length):
        for header, data in data_by_header.items():
            if i < len(data):
                text_widget.insert(tk.END, f"{str(data[i]).ljust(12)}")
            else:
                text_widget.insert(tk.END, "".ljust(12))
        text_widget.insert(tk.END, '\n')

def crear_producto(text_widget):
    book = openpyxl.load_workbook('inventario.xlsx')
    sheet = book['Inventario']

    producto = simpledialog.askinteger("Input", "Ingrese el código del producto:")
    nombre = simpledialog.askstring("Input", "Ingrese el nombre del producto:")
    existencia = simpledialog.askinteger("Input", "Ingrese la existencia del producto:")
    proveedor = simpledialog.askstring("Input", "Ingrese el proveedor del producto:")
    precio = simpledialog.askstring("Input", "Ingrese el precio del producto:")

    data = [producto, nombre, existencia, proveedor, precio]
    sheet.append(data)

    book.save('inventario.xlsx')

def actualizar_producto(text_widget):
    try:
        book = openpyxl.load_workbook('inventario.xlsx')
        sheet = book['Inventario']

        codigo = simpledialog.askinteger("Input", "Ingrese el código del producto que desea actualizar:")

        producto_encontrado = False

        for fila in sheet.iter_rows(min_col=1, max_col=1):
            celda = fila[0]
            if celda.value == codigo:
                text_widget.insert(tk.END, "Se ha encontrado el producto\n")
                producto_encontrado = True
                fila_destino = celda.row 
                columna_destino = 5
                celda_destino = sheet.cell(row=fila_destino, column=columna_destino)
                precio_nuevo = simpledialog.askstring("Input", "Ingrese el nuevo precio:")
                celda_destino.value = precio_nuevo
                break 

        if not producto_encontrado:
            text_widget.insert(tk.END, "No se encontró el producto con el código proporcionado\n")

        book.save('inventario.xlsx')

    except Exception as e:
        text_widget.insert(tk.END, f"Se produjo un error: {str(e)}\n")

def actualizar_existencia(text_widget):
    try:
        book = openpyxl.load_workbook('inventario.xlsx')
        sheet = book['Inventario']

        codigo = simpledialog.askinteger("Input", "Ingrese el código del producto que desea actualizar:")

        producto_encontrado = False

        for fila in sheet.iter_rows(min_col=1, max_col=1):
            celda = fila[0]
            if celda.value == codigo:
                text_widget.insert(tk.END, "Se ha encontrado el producto\n")
                producto_encontrado = True
                fila_destino = celda.row 
                columna_destino = 3
                celda_destino = sheet.cell(row=fila_destino, column=columna_destino)
                precio_nuevo = simpledialog.askstring("Input", "Ingrese la nueva existencia:")
                celda_destino.value = precio_nuevo
                break 

        if not producto_encontrado:
            text_widget.insert(tk.END, "No se encontró el producto con el código proporcionado\n")

        book.save('inventario.xlsx')

    except Exception as e:
        text_widget.insert(tk.END, f"Se produjo un error: {str(e)}\n")

def eliminar_producto(text_widget):
    try:
        book = openpyxl.load_workbook('inventario.xlsx')
        sheet = book['Inventario']

        codigo = simpledialog.askinteger("Input", "Ingrese el código del producto que desea eliminar:")

        producto_encontrado = False

        for fila in sheet.iter_rows(min_col=1, max_col=1):
            celda = fila[0]
            if celda.value == codigo:
                text_widget.insert(tk.END, "Se ha encontrado el producto\n")
                producto_encontrado = True
                fila_destino = celda.row 
                sheet.delete_rows(fila_destino)
                text_widget.insert(tk.END, "El producto ha sido eliminado.\n")
                break 

        if not producto_encontrado:
            text_widget.insert(tk.END, "No se encontró el producto con el código proporcionado\n")

        book.save('inventario.xlsx')

    except Exception as e:
        text_widget.insert(tk.END, f"Se produjo un error: {str(e)}\n")

#Modulo Clientes

def listar_clientes(text_widget):

    text_widget.delete('1.0', tk.END)

    book = openpyxl.load_workbook('inventario.xlsx')
    sheet = book['Clientes']

    headers = [cell.value for cell in sheet[1]]

    data_by_header = {header: [] for header in headers}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        for header, cell_value in zip(headers, row):
            data_by_header[header].append(cell_value)

    max_data_length = max(len(data) for data in data_by_header.values())

    for header in headers:
        text_widget.insert(tk.END, f"{header.ljust(12)}")
    text_widget.insert(tk.END, '\n')

    for i in range(max_data_length):
        for header, data in data_by_header.items():
            if i < len(data):
                text_widget.insert(tk.END, f"{str(data[i]).ljust(12)}")
            else:
                text_widget.insert(tk.END, "".ljust(12))
        text_widget.insert(tk.END, '\n')


def crear_clientes(text_widget):
    book = openpyxl.load_workbook('inventario.xlsx')
    sheet = book['Clientes']

    while True:
        codigo = simpledialog.askinteger("Input", "Ingrese el código del cliente a agregar: ")
        
        if codigo == -1:
            break

        nombre = simpledialog.askstring("Input", "Ingrese el nombre del cliente:")
        Direccion = simpledialog.askstring("Input", "Ingrese la dirección del cliente:")

        clientes_existentes = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            codigo_cliente, nombre_cliente, direccion_cliente = row
            clientes_existentes.append((codigo_cliente, nombre_cliente, direccion_cliente))

        if any(cliente[0] == codigo for cliente in clientes_existentes):
            text_widget.insert(tk.END, f"El cliente con código {codigo} ya existe. Introduzca otro código.\n")
        else:
            nuevo_cliente = (codigo, nombre, Direccion)
            clientes_existentes.append(nuevo_cliente)

            clientes_ordenados = sorted(clientes_existentes, key=lambda cliente: cliente[0])

            for _ in range(2, sheet.max_row + 1):
                sheet.delete_rows(2)

            for cliente in clientes_ordenados:
                sheet.append(cliente)

            text_widget.insert(tk.END, "El nuevo cliente fue agregado de forma correcta\n")

            book.save('inventario.xlsx')
            break

def actualizar_cliente(text_widget):
    try:
        book = openpyxl.load_workbook('inventario.xlsx')
        sheet = book['Clientes']

        codigo = simpledialog.askinteger("Input", "Ingrese el código del cliente que desea actualizar:")

        cliente_encontrado = False

        for fila in sheet.iter_rows(min_col=1, max_col=1):
            celda = fila[0]
            if celda.value == codigo:
                text_widget.insert(tk.END, "Se ha encontrado el cliente\n")
                cliente_encontrado = True
                fila_destino = celda.row 
                columna_destino = 3
                celda_destino = sheet.cell(row=fila_destino, column=columna_destino)
                direccion_nueva = simpledialog.askstring("Input", "Ingrese la nueva dirección:")
                celda_destino.value = direccion_nueva
                text_widget.insert(tk.END, "Los datos del cliente se actualizaron correctamente\n")
                break 

        if not cliente_encontrado:
            text_widget.insert(tk.END, "No se encontró el cliente con el código proporcionado\n")

        book.save('inventario.xlsx')

    except Exception as e:
        text_widget.insert(tk.END, f"Se produjo un error: {str(e)}\n")

def eliminar_cliente(text_widget):
    try:
        book = openpyxl.load_workbook('inventario.xlsx')
        sheet = book['Clientes']

        codigo = simpledialog.askinteger("Input", "Ingrese el código del cliente que desea eliminar:")

        cliente_encontrado = False

        for fila in sheet.iter_rows(min_col=1, max_col=1):
            celda = fila[0]
            if celda.value == codigo:
                text_widget.insert(tk.END, "Se ha encontrado el cliente\n")
                cliente_encontrado = True
                fila_destino = celda.row 
                sheet.delete_rows(fila_destino) 
                text_widget.insert(tk.END, "El cliente ha sido eliminado.\n")
                break 

        if not cliente_encontrado:
            text_widget.insert(tk.END, "No se encontró el cliente con el código proporcionado\n")

        book.save('inventario.xlsx')

    except Exception as e:
        text_widget.insert(tk.END, f"Se produjo un error: {str(e)}\n")


#Modulo de Ventas

def listar_ventas(text_widget):

    text_widget.delete('1.0', tk.END)

    book = openpyxl.load_workbook('inventario.xlsx')
    sheet = book['Ventas']

    headers = [cell.value for cell in sheet[1]]

    data_by_header = {header: [] for header in headers}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        for header, cell_value in zip(headers, row):
            data_by_header[header].append(cell_value)

    max_data_length = max(len(data) for data in data_by_header.values())

    for header in headers:
        text_widget.insert(tk.END, f"{header.ljust(14)}")
    text_widget.insert(tk.END, '\n')

    for i in range(max_data_length):
        for header, data in data_by_header.items():
            if i < len(data):
                text_widget.insert(tk.END, f"{str(data[i]).ljust(20)}")
            else:
                text_widget.insert(tk.END, "".ljust(12))
        text_widget.insert(tk.END, '\n')


def verificar_producto_en_inventario(codigo_producto, hoja_inventario):
    for row in hoja_inventario.iter_rows(values_only=True):
        if row[0] == codigo_producto:
            return row

    return None

def agregar_venta(text_widget):
    codigo_producto = simpledialog.askinteger("Input", "Ingrese el código del producto a vender:")
    codigo_cliente = simpledialog.askinteger("Input", "Ingrese el código del cliente:")

    book = openpyxl.load_workbook('inventario.xlsx')
    
    hoja_inventario = book['Inventario']

    producto = verificar_producto_en_inventario(codigo_producto, hoja_inventario)

    if producto is None:
        text_widget.insert(tk.END, f"El producto con código {codigo_producto} no se encuentra en el inventario.\n")
        return

    cantidad_existente = producto[2] 
    precio_unitario = float(producto[4])  

    if cantidad_existente <= 0:
        text_widget.insert(tk.END, "El producto está agotado y no se puede vender.\n")
        return

    cantidad_vendida = simpledialog.askinteger("Input", f"Ingrese la cantidad a vender (existencia actual: {cantidad_existente}):")

    if cantidad_vendida > cantidad_existente:
        text_widget.insert(tk.END, "No hay suficiente cantidad en inventario para la venta.\n")
        return

    total_venta = cantidad_vendida * precio_unitario

    hoja_ventas = book['Ventas']
    hoja_ventas.append([codigo_producto, codigo_cliente, cantidad_vendida, total_venta])

    for idx, row in enumerate(hoja_inventario.iter_rows(values_only=True, min_row=2), start=2):
        if row[0] == codigo_producto:
            cantidad_actual = row[2]
            nueva_cantidad = cantidad_actual - cantidad_vendida
            hoja_inventario.cell(row=idx, column=3, value=nueva_cantidad)

    book.save('inventario.xlsx')

    text_widget.insert(tk.END, f"Venta registrada exitosamente. Total de la venta: {total_venta}.\n")

def anular_venta(text_widget):
    book = openpyxl.load_workbook('inventario.xlsx')
    sheet = book['Ventas']

    codigo_producto = simpledialog.askinteger("Input", "Ingrese el codigo del producto vendido, por medio de este se eliminará la venta:")
    codigo_cliente = simpledialog.askinteger("Input", "Ingrese el codigo del cliente que realizó la venta:")

    venta_encontrada = False

    for fila in sheet.iter_rows(min_row=2):
        celda_producto = fila[0]
        celda_cliente = fila[1]
        if celda_producto.value == codigo_producto and celda_cliente.value == codigo_cliente:
            text_widget.insert(tk.END, "Se ha encontrado la venta\n")
            venta_encontrada = True
            fila_destino = celda_producto.row 
            sheet.delete_rows(fila_destino) 
            text_widget.insert(tk.END, "la venta fue anulada/eliminada.\n")
            break 

    if not venta_encontrada:
        text_widget.insert(tk.END, "No se encontró la venta con el código de producto y cliente proporcionados\n")

    book.save('inventario.xlsx')


#Modulo consultas

def generar_informe_ventas_por_producto(text_widget):
    archivo_excel = 'inventario.xlsx'
    hoja_ventas = 'Ventas'
    nombre_archivo_salida = 'informe_ventas_por_producto.xlsx'

    try:
        book = openpyxl.load_workbook(archivo_excel)
        sheet = book[hoja_ventas]

        ventas_por_producto = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            codigo_producto, _, cantidad_producto, total_venta = row 

            
            if codigo_producto in ventas_por_producto:
                ventas_por_producto[codigo_producto]["Cantidad Vendida"] += cantidad_producto
                ventas_por_producto[codigo_producto]["Total Ventas"] += total_venta
            else:
               
                ventas_por_producto[codigo_producto] = {
                    "Cantidad Vendida": cantidad_producto,
                    "Total Ventas": total_venta
                }

        informe_book = openpyxl.Workbook()
        informe_sheet = informe_book.active
        informe_sheet.title = "Informe Ventas por Producto"

        informe_sheet.append(["Código de Producto", "Cantidad Total Vendida", "Total Ventas"])

        for codigo_producto, datos in ventas_por_producto.items():
            informe_sheet.append([codigo_producto, datos["Cantidad Vendida"], datos["Total Ventas"]])

        informe_book.save(nombre_archivo_salida)
        text_widget.insert(tk.END, f"Informe de ventas por producto generado en '{nombre_archivo_salida}'.\n")

    except Exception as e:
        text_widget.insert(tk.END, f"Se produjo un error: {str(e)}\n")

def generar_informe_ventas_por_cliente(text_widget):
    archivo_excel = 'inventario.xlsx'
    hoja_ventas = 'Ventas'
    nombre_archivo_salida = 'informe_ventas_por_cliente.xlsx'

    book = openpyxl.load_workbook(archivo_excel)
    sheet = book[hoja_ventas]

    ventas_por_cliente = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        codigo_producto, codigo_cliente, cantidad_producto, total_venta = row

        codigo_cliente_str = str(codigo_cliente)

        total_gastado = total_venta

        if codigo_cliente_str in ventas_por_cliente:
            ventas_por_cliente[codigo_cliente_str] += total_gastado
        else:
            ventas_por_cliente[codigo_cliente_str] = total_gastado

    nuevo_book = openpyxl.Workbook()
    nueva_sheet = nuevo_book.active

    nueva_sheet['A1'] = "Código del Cliente"
    nueva_sheet['B1'] = "Total Gastado"

    fila = 2
    for codigo_cliente, total_gastado in ventas_por_cliente.items():
        nueva_sheet[f'A{fila}'] = int(codigo_cliente) 
        nueva_sheet[f'B{fila}'] = total_gastado
        fila += 1

    nuevo_book.save(nombre_archivo_salida)

    text_widget.insert(tk.END, f"Informe de ventas por cliente generado en '{nombre_archivo_salida}'.\n")

#Modulo gmail

def enviar_correo(text_widget):
    servidor_smtp = 'smtp.gmail.com'  
    puerto = 587  
    usuario = 'dereksalguero236@gmail.com'
    contrasena = 'ppqi grft nayi qdef'
    destinatario = simpledialog.askstring("Input", "Ingrese el correo electrónico del destinatario:")
    asunto = 'Informe de ventas'
    cuerpo = 'Aquí está el informe de ventas que solicitaste.'
    archivo = filedialog.askopenfilename(title="Seleccione el archivo a enviar")

    msg = MIMEMultipart()
    msg['From'] = usuario
    msg['To'] = destinatario
    msg['Subject'] = asunto

    msg.attach(MIMEText(cuerpo, 'plain'))

    adjunto = open(archivo, 'rb')

    part = MIMEBase('application', 'octet-stream')
    part.set_payload((adjunto).read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename= ' + archivo)

    msg.attach(part)

    servidor = smtplib.SMTP(servidor_smtp, puerto)
    servidor.starttls()
    servidor.login(usuario, contrasena)
    text = msg.as_string()
    servidor.sendmail(usuario, destinatario, text)
    servidor.quit()


def main():
    root = tk.Tk()
    root.title("Inventario")

    frame1 = tk.Frame(root)
    frame1.pack()

    frame2 = tk.Frame(root)
    frame2.pack()

    text_widget = tk.Text(frame2)
    text_widget.pack()

    # Crea un menú desplegable
    menu = tk.Menu(root)
    root.config(menu=menu)

    # Crea el menú de inventario y añádelo al menú desplegable
    inventario_menu = tk.Menu(menu)
    menu.add_cascade(label="Inventario", menu=inventario_menu)

    #Botones menú inventario
    inventario_menu.add_command(label="Listar Productos", command=lambda: listar_producto(text_widget))
    inventario_menu.add_command(label="Crear Producto", command=lambda: crear_producto(text_widget))
    inventario_menu.add_command(label="Actualizar Producto", command=lambda: actualizar_producto(text_widget))
    inventario_menu.add_command(label="Actualizar Existencia", command=lambda: actualizar_existencia(text_widget))
    inventario_menu.add_command(label="Eliminar Producto", command=lambda: eliminar_producto(text_widget))

    #Menú clientes
    clientes_menu = tk.Menu(menu)
    menu.add_cascade(label="Clientes", menu=clientes_menu)

    #Botones menú clientes
    clientes_menu.add_command(label="Listar Clientes", command=lambda: listar_clientes(text_widget))
    clientes_menu.add_command(label="Crear Cliente", command=lambda: crear_clientes(text_widget))
    clientes_menu.add_command(label="Actualizar Cliente", command=lambda: actualizar_cliente(text_widget))
    clientes_menu.add_command(label="Eliminar Cliente", command=lambda: eliminar_cliente(text_widget))

    #Menu Ventas
    ventas_menu = tk.Menu(menu)
    menu.add_cascade(label="Ventas", menu=ventas_menu)

    #Botones menú ventas
    ventas_menu.add_command(label="Listar Ventas", command=lambda: listar_ventas(text_widget))
    ventas_menu.add_command(label="Agregar Venta", command=lambda: agregar_venta(text_widget))
    ventas_menu.add_command(label="Anular venta", command=lambda: anular_venta(text_widget))

    #Menu Consultas
    consultas_menu = tk.Menu(menu)
    menu.add_cascade(label="Consultas", menu=consultas_menu)

    #Botones menú consulta
    consultas_menu.add_command(label="Generar informe de clientes", command=lambda: generar_informe_ventas_por_cliente(text_widget))
    consultas_menu.add_command(label="Generar informe de producto", command=lambda: generar_informe_ventas_por_producto(text_widget))

    menu.add_command(label="Enviar informe de ventas por correo", command=lambda: enviar_correo(text_widget))

    root.mainloop()

if __name__ == "__main__":
    main()